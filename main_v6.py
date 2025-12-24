# =============================================
# BEAST MODE LINKEDIN SCRAPER (Daily Excel + Append + Dedupe)
# Console Output + Excel + Last 7 Days + Sorted
# =============================================

import asyncio
import httpx
from bs4 import BeautifulSoup
import re
import random
import os
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook

# ---------------------------------------------
# CONFIG
# ---------------------------------------------
KEYWORDS = ["Entry Level 2026", "Graduate Engineer"]
LOCATION = "United States"
MAX_RETRY = 3

# ---------------------------------------------
# DATE FILTER (Last 7 days)
# ---------------------------------------------
def posted_within_last_week(date_str):
    try:
        post_date = datetime.fromisoformat(date_str.replace("Z", ""))
        return post_date >= datetime.now() - timedelta(days=7)
    except:
        return False

# ---------------------------------------------
# CLEAN JOB ID → Mobile Link
# ---------------------------------------------
def get_mobile_link(job_link):
    match = re.search(r"/jobs/view/.*?-(\d+)", job_link)
    return f"https://www.linkedin.com/jobs/view/{match.group(1)}" if match else job_link

# ---------------------------------------------
# SCRAPE PER KEYWORD
# ---------------------------------------------
async def fetch_jobs_for_keyword(client, keyword):
    url = "https://www.linkedin.com/jobs-guest/jobs/api/seeMoreJobPostings/search"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                      "(KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36"
    }

    job_postings = []
    seen_ids_today = set()

    print(f"\n🚀 Starting scrape for keyword: {keyword}")
    page = 0
    retry = 0

    while True:
        params = {
            "keywords": keyword,
            "location": LOCATION,
            "sortBy": "R",          # Most recent
            "f_TPR": "r86400",     # Last 7 days
            "start": page * 25
        }

        print(f"🔄 Fetching page {page}...")

        try:
            resp = await client.get(url, headers=headers, params=params)

            # RATE LIMIT
            if resp.status_code == 429:
                wait_time = random.randint(45, 90)
                print(f"🚫 Rate limited (429). Sleeping {wait_time} sec...")
                await asyncio.sleep(wait_time)
                continue

            # OTHER ERRORS
            if resp.status_code != 200:
                print(f"❌ HTTP {resp.status_code} on page {page}")
                retry += 1
                if retry > MAX_RETRY:
                    print("❌ Too many failures. Stopping scraper.")
                    break
                await asyncio.sleep(5)
                continue

            retry = 0
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(resp.content, "lxml")
            job_cards = soup.select("li")

            if not job_cards:
                print(f"ℹ️ No more jobs found for: {keyword}")
                break

            page += 1

            for job in job_cards:
                try:
                    title_tag = job.find("h3")
                    company_tag = job.find("h4")
                    location_tag = job.find("span", class_="job-search-card__location")
                    time_tag = job.find("time")
                    link_tag = job.find("a", href=True)

                    if not all([title_tag, company_tag, location_tag, time_tag, link_tag]):
                        continue

                    job_link = link_tag["href"]
                    mobile_link = get_mobile_link(job_link)

                    job_id_match = re.search(r"/jobs/view/.*?-(\d+)", job_link)
                    if not job_id_match:
                        continue
                    job_id = job_id_match.group(1)

                    if job_id in seen_ids_today:
                        continue
                    seen_ids_today.add(job_id)

                    title = title_tag.text.strip()
                    company = company_tag.text.strip()
                    location = location_tag.text.strip()
                    date_posted = time_tag.get("datetime", "")

                    if not posted_within_last_week(date_posted):
                        continue

                    job_postings.append({
                        "Job ID": job_id,
                        "Title": title,
                        "Company": company,
                        "Location": location,
                        "Date Posted": date_posted,
                        "Keyword": keyword,
                        "Job Link": job_link,
                        "Mobile Link": mobile_link
                    })

                except Exception as e:
                    print(f"⚠️ Parse Error: {e}")
                    continue

            await asyncio.sleep(random.uniform(1.5, 3.5))

        except Exception as e:
            print(f"⚠️ Fatal Error: {e}")
            break

    print(f"✅ Completed '{keyword}' — {len(job_postings)} jobs found.\n")
    return job_postings

# ---------------------------------------------
# CLEAN SHEET NAME
# ---------------------------------------------
def clean_sheet_name(name):
    return re.sub(r"[\\/*?:\[\]]", "", name.replace(" ", "_"))

# ---------------------------------------------
# MAIN
# ---------------------------------------------
async def main():
    date_code = datetime.now().strftime("%Y%m%d")
    excel_file = f"LinkedIn_Jobs_{date_code}.xlsx"

    # Load or create workbook
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        print(f"📁 Loaded existing Excel: {excel_file}")
    else:
        wb = Workbook()
        print(f"📁 Creating new Excel: {excel_file}")

    # Delete default sheet if only exists and empty
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        ws_default = wb["Sheet"]
        wb.remove(ws_default)

    async with httpx.AsyncClient(timeout=45.0) as client:
        for keyword in KEYWORDS:
            jobs = await fetch_jobs_for_keyword(client, keyword)
            if not jobs:
                print(f"ℹ️ No jobs for {keyword}, skipping sheet creation.")
                continue

            sheet_name = clean_sheet_name(keyword)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"📌 Sheet exists. Appending to existing sheet: {sheet_name}")
                # Load existing job IDs
                existing_ids = set()
                for row in ws.iter_rows(min_row=2, values_only=True):
                    existing_ids.add(str(row[0]))
            else:
                ws = wb.create_sheet(title=sheet_name)
                ws.append(["Job ID", "Title", "Company", "Location", "Date Posted",
                           "Keyword", "Job Link", "Mobile Link"])
                existing_ids = set()

            # Append new jobs, skip duplicates
            new_count = 0
            for job in jobs:
                if job["Job ID"] in existing_ids:
                    continue
                ws.append([job["Job ID"], job["Title"], job["Company"], job["Location"],
                           job["Date Posted"], job["Keyword"], job["Job Link"], job["Mobile Link"]])
                existing_ids.add(job["Job ID"])
                new_count += 1

            print(f"📝 {new_count} new jobs added to sheet '{sheet_name}'.")

            # Console output
            print("\n========================")
            print(f"📢 RESULTS FOR {keyword}")
            print("========================\n")
            for job in jobs:
                print("🧾 Title   :", job["Title"])
                print("🏢 Company :", job["Company"])
                print("📍 Location:", job["Location"])
                print("📅 Posted  :", job["Date Posted"])
                print("🔑 Keyword :", job["Keyword"])
                print("🔗 Link    :", job["Mobile Link"])
                print("-" * 80)

    # Save workbook
    wb.save(excel_file)
    print(f"\n🎉 Excel saved: {excel_file}")
    print("✅ Scraping Completed Successfully!")

# ---------------------------------------------
# RUN
# ---------------------------------------------
if __name__ == "__main__":
    asyncio.run(main())
