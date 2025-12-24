# =============================================
#   BEAST MODE LINKEDIN SCRAPER (Final)
#   Excel daily file + per-keyword sheets + dedupe
# =============================================

import asyncio
import httpx
from bs4 import BeautifulSoup
import re
import random
import os
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook

# Excel support
# try:
#     from openpyxl import Workbook, load_workbook
# except ImportError:
#     print("Missing dependency: openpyxl. Install with: pip install openpyxl")
#     raise

# ---------------------------------------------
# CONFIG
# ---------------------------------------------
KEYWORDS = ["Entry Level 2026","Graduate Engineer"]         # add more keywords as needed
LOCATION = "United States"
MAX_RETRY = 3

# ---------------------------------------------
# DATE FILTER (Last 7 days)
# ---------------------------------------------
def posted_within_last_week(date_str):
    """
    Expecting ISO datetime in the datetime attribute of <time>.
    Returns True if posted within last 7 days.
    """
    try:
        post_date = datetime.fromisoformat(date_str.replace("Z", ""))
        return post_date >= datetime.now() - timedelta(days=7)
    except Exception:
        return False


# ---------------------------------------------
# CLEAN JOB ID → Mobile Link
# ---------------------------------------------
def get_mobile_link(job_link):
    match = re.search(r"/jobs/view/.*?-(\d+)", job_link)
    return f"https://www.linkedin.com/jobs/view/{match.group(1)}" if match else job_link


# ---------------------------------------------
# LOAD PREVIOUS DAY JOB IDS FOR DEDUPLICATION (from yesterday's Excel)
# ---------------------------------------------
def load_previous_ids(keyword):
    """
    Reads yesterday's Excel file (Job_Extract_YYYYMMDD.xlsx).
    Returns set of job ids (strings) found in the sheet for the keyword.
    """
    yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y%m%d")
    file = f"Job_Extract_{yesterday}.xlsx"
    safe_kw = keyword.replace(" ", "")

    if not os.path.exists(file):
        return set()

    ids = set()
    try:
        wb = load_workbook(file, read_only=True, data_only=True)
        if safe_kw not in wb.sheetnames:
            return set()
        ws = wb[safe_kw]

        # Assume columns: Title, Company, Location, Date Posted, Keyword, Job Link, Mobile Link
        # Job Link is column index 6 (0-based -> 5)
        for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
            link = ""
            # handle if fewer columns present
            try:
                link = row[5]  # job link column (0-based index 5)
            except Exception:
                link = None
            if not link:
                continue
            m = re.search(r"/jobs/view/.*?-(\d+)", str(link))
            if m:
                ids.add(m.group(1))
    except Exception:
        # If reading fails for any reason, return empty set to avoid blocking scraping
        return set()

    return ids


# ---------------------------------------------
# SCRAPE PER KEYWORD
# ---------------------------------------------
async def fetch_jobs_for_keyword(client, keyword):
    """
    Fetch jobs for a single keyword and return list of job dicts.
    Uses yesterday's sheet to dedupe cross-day.
    """
    url = "https://www.linkedin.com/jobs-guest/jobs/api/seeMoreJobPostings/search"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                      "(KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36"
    }

    job_postings = []
    seen_ids_today = set()
    seen_ids_yesterday = load_previous_ids(keyword)

    print(f"\n🚀 Starting scrape for keyword: {keyword}")
    print(f"📌 Loaded {len(seen_ids_yesterday)} previous job IDs (for dedupe)")

    page = 0
    retry = 0

    while True:
        params = {
            "keywords": keyword,
            "location": LOCATION,
            "sortBy": "R",          # Most recent
            "f_TPR": "r86400",     # Last 7 days
            "start": page * 25
        }

        print(f"🔄 Fetching page {page}...")

        try:
            resp = await client.get(url, headers=headers, params=params)

            # RATE LIMIT
            if resp.status_code == 429:
                wait_time = random.randint(45, 90)
                print(f"🚫 Rate limited (429). Sleeping {wait_time} sec...")
                await asyncio.sleep(wait_time)
                continue

            # OTHER ERRORS
            if resp.status_code != 200:
                print(f"❌ HTTP {resp.status_code} on page {page}")
                retry += 1
                if retry > MAX_RETRY:
                    print("❌ Too many failures. Stopping scraper.")
                    break
                await asyncio.sleep(5)
                continue

            retry = 0
            soup = BeautifulSoup(resp.content, "lxml")
            job_cards = soup.select("li")

            if not job_cards:
                print(f"ℹ️ No more jobs found for: {keyword}")
                break

            page += 1

            for job in job_cards:
                try:
                    title_tag = job.find("h3")
                    company_tag = job.find("h4")
                    location_tag = job.find("span", class_="job-search-card__location")
                    time_tag = job.find("time")
                    link_tag = job.find("a", href=True)

                    if not all([title_tag, company_tag, location_tag, time_tag, link_tag]):
                        continue

                    job_link = link_tag["href"]
                    mobile_link = get_mobile_link(job_link)

                    job_id_match = re.search(r"/jobs/view/.*?-(\d+)", job_link)
                    if not job_id_match:
                        continue
                    job_id = job_id_match.group(1)

                    # Dedupe: today's run
                    if job_id in seen_ids_today:
                        continue

                    # Dedupe: yesterday's run (main requirement)
                    if job_id in seen_ids_yesterday:
                        continue

                    # Mark seen for today's run
                    seen_ids_today.add(job_id)

                    title = title_tag.text.strip()
                    company = company_tag.text.strip()
                    location = location_tag.text.strip()
                    date_posted = time_tag.get("datetime", "")

                    # Ensure last 7 days filter
                    if not posted_within_last_week(date_posted):
                        continue

                    job_postings.append({
                        "Title": title,
                        "Company": company,
                        "Location": location,
                        "Date Posted": date_posted,
                        "Keyword": keyword,
                        "Job Link": job_link,
                        "Mobile Link": mobile_link
                    })

                except Exception as e:
                    print(f"⚠️ Parse Error: {e}")
                    continue

            # polite randomized delay
            await asyncio.sleep(random.uniform(1.5, 3.5))

        except Exception as e:
            print(f"⚠️ Fatal Error: {e}")
            break

    print(f"✅ Completed '{keyword}' — {len(job_postings)} jobs found.\n")
    return job_postings


# ---------------------------------------------
# MAIN: orchestrates scraping and Excel writing
# ---------------------------------------------
async def main():
    # Prepare today's excel file
    date_code = datetime.now().strftime("%Y%m%d")
    excel_file = f"Job_Extract_{date_code}.xlsx"

    # Create workbook for today if not exists
    if not os.path.exists(excel_file):
        wb_new = Workbook()
        # remove the default sheet that openpyxl creates
        default = wb_new.active
        wb_new.remove(default)
        # ---- FIX: Ensure at least one visible sheet exists ----
        if not wb_new.sheetnames:
            wb_new.create_sheet("Sheet1")

        # Make sure first sheet is visible
        first_sheet = wb_new[wb_new.sheetnames[0]]
        first_sheet.sheet_state = "visible"
        # --------------------------------------------------------

        wb_new.save(excel_file)

    async with httpx.AsyncClient(timeout=45.0) as client:
        # load workbook (read/write)
        wb = load_workbook(excel_file)
        for keyword in KEYWORDS:
            jobs = await fetch_jobs_for_keyword(client, keyword)

            # Prepare sheet name & remove if exists (we want latest results on today's sheet)
            safe_kw = keyword.replace(" ", "")
            if safe_kw in wb.sheetnames:
                # remove to overwrite fresh results
                ws_old = wb[safe_kw]
                wb.remove(ws_old)
            ws = wb.create_sheet(safe_kw)

            # Header row
            headers = ["Title", "Company", "Location", "Date Posted",
                       "Keyword", "Job Link", "Mobile Link"]
            ws.append(headers)

            # Append job rows
            for job in jobs:
                ws.append([
                    job.get("Title", ""),
                    job.get("Company", ""),
                    job.get("Location", ""),
                    job.get("Date Posted", ""),
                    job.get("Keyword", ""),
                    job.get("Job Link", ""),
                    job.get("Mobile Link", "")
                ])

            # Console output
            print("\n========================")
            print(f"📢 RESULTS FOR {keyword}")
            print("========================\n")
            for job in jobs:
                print("🧾 Title   :", job["Title"])
                print("🏢 Company :", job["Company"])
                print("📍 Location:", job["Location"])
                print("📅 Posted  :", job["Date Posted"])
                print("🔑 Keyword :", job["Keyword"])
                print("🔗 Link    :", job["Mobile Link"])
                print("-" * 80)

        # Save workbook once after all keywords processed
        wb.save(excel_file)
        print(f"\n📁 Saved to Excel File: {excel_file}")
        print("🎉 Scraping Completed Successfully!\n")

        # DELETE DEFAULT SHEET SAFELY
        if "Sheet" in wb.sheetnames:
            ws_default = wb["Sheet"]
            wb.remove(ws_default)

# ---------------------------------------------
# RUN
# ---------------------------------------------
if __name__ == "__main__":
    asyncio.run(main())
