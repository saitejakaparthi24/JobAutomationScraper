import time
import random
import requests
from openpyxl import load_workbook
from datetime import datetime

# ====================================
# WHATSAPP CLOUD API CONFIG
# ====================================
ACCESS_TOKEN = "EAAMBPp8HmjwBQOL6760PbaBXFG1rZAmrqWyVzEiAWuGgqjA68cHZBNkJ2zOsS3qixqbFVnn7473kNqybCUYRUfaWfih3iXKAJVLegJgO9NFoiXULKc3B9EbZBncWyagT2NWBoTgjIrq3doTAL8g2pTgczxRH9cDopmts41OMCuqqJ2Q1Xj4gAF1KqMFe33WJeFbEf72FfpxyfPcrI8g4ZBbwAmZAlK0QurShGDB0QERVaJu63CLBeUpv0AcIphEkGlHYEGgGe9Gs3sF6ZBFZAlS1SZB8askdQkIvVnsLZBAZDZD"
PHONE_NUMBER_ID = "902611009600754"
RECIPIENT_NUMBER = "+18064518285"   # Replace with recipient number

# ====================================
# LOAD TODAY'S EXCEL FILE
# ====================================
today_code = datetime.now().strftime("%Y%m%d")
excel_file = f"Job_Extract_{today_code}.xlsx"

def load_all_jobs(excel_path):
    wb = load_workbook(excel_path)
    all_jobs = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            title, company, location, date_posted, keyword, link, mobile = row
            all_jobs.append({
                "title": title or "",
                "company": company or "",
                "location": location or "",
                "date": str(date_posted) if date_posted else "",
                "keyword": keyword or "",
                "link": mobile or ""
            })
    return all_jobs

# ====================================
# SEND MESSAGE TO WHATSAPP
# ====================================
def send_whatsapp_message(message):
    url = f"https://graph.facebook.com/v22.0/{PHONE_NUMBER_ID}/messages"
    payload = {
        "messaging_product": "whatsapp",
        "to": RECIPIENT_NUMBER,
        "type": "text",
        "text": {"body": message}
    }
    headers = {
        "Authorization": f"Bearer {ACCESS_TOKEN}",
        "Content-Type": "application/json"
    }
    try:
        response = requests.post(url, json=payload, headers=headers)
        print("📨 WhatsApp Response:", response.json())
        return response.json()
    except Exception as e:
        print(f"⚠️ WhatsApp Send Error: {e}")
        return None

# ====================================
# CREATE MESSAGE FROM JOBS
# ====================================
def create_message_from_jobs(jobs):
    text = "🔥 *Top Jobs for You* 🔥\n\n"
    for j in jobs:
        text += f"📌 *{j['title']}*\n"
        text += f"🏢 {j['company']}\n"
        text += f"📍 {j['location']}\n"
        text += f"🔗 {j['link']}\n"
        text += "--------------------\n"
    return text

# ====================================
# MAIN LOOP — SEND BATCHES EVERY 30 MINUTES
# ====================================
def start_auto_sender(batch_size=2, interval_minutes=1):
    all_jobs = load_all_jobs(excel_file)
    print(f"📥 Loaded {len(all_jobs)} jobs from Excel.")

    sent_jobs = set()  # Track sent jobs

    while all_jobs:
        available_jobs = [j for j in all_jobs if j['link'] not in sent_jobs]
        if not available_jobs:
            print("✅ All jobs have been sent. Exiting.")
            break

        batch = random.sample(available_jobs, min(batch_size, len(available_jobs)))
        message = create_message_from_jobs(batch)
        send_whatsapp_message(message)

        for j in batch:
            sent_jobs.add(j['link'])

        print(f"⏳ Waiting {interval_minutes} minutes before next batch...\n")
        time.sleep(interval_minutes * 60)

# ====================================
# START
# ====================================
if __name__ == "__main__":
    start_auto_sender(batch_size=2, interval_minutes=1)
